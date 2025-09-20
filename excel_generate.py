import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ---------- Helper functions ----------
def extract_key_from_path(path: str) -> str:
    """
    Extract the keyword inside getter/is methods from a full path.
    Example: abc.def.getHai12() -> hai12
    """
    match = re.search(r'\b(?:get|is)([A-Za-z0-9_]+)\s*\(\)', path)
    if match:
        return match.group(1).lower()  # case-insensitive
    return None

def normalize_field(field: str) -> str:
    """
    Normalize fields by removing get/is and parentheses.
    Example: getCustomerName() -> customername
    """
    field = field.strip()
    match = re.match(r'(?:get|is)([A-Za-z0-9_]+)\s*\(\)', field, re.IGNORECASE)
    if match:
        return match.group(1).lower()
    return field.lower()

# ---------- Load JSON files ----------
with open("master_dictionary.json", "r") as f:
    master_dict = json.load(f)

with open("output.json", "r") as f:
    output_dict = json.load(f)

# ---------- Process Matching ----------
matched_rows = []
unmatched_paths = []

for path, master_fields in master_dict.items():
    extracted_key = extract_key_from_path(path)
    if not extracted_key:
        continue

    matched = False
    for out_key, out_fields in output_dict.items():
        if extracted_key.lower() == out_key.lower():
            matched = True

            master_fields_norm = [normalize_field(f) for f in master_fields]
            out_fields_norm = [normalize_field(f) for f in out_fields]

            matched_fields = list(set(master_fields_norm) & set(out_fields_norm))
            unmatched_master = list(set(master_fields_norm) - set(out_fields_norm))
            unmatched_output = list(set(out_fields_norm) - set(master_fields_norm))

            matched_rows.append({
                "path": path,
                "matchedkey": out_key,
                "matchedfields": ", ".join(matched_fields),
                "matchedcount": len(matched_fields),
                "unmatchedfields": ", ".join(unmatched_master + unmatched_output),
                "unmatchedcount": len(unmatched_master) + len(unmatched_output),
                "present_in_jar": ", ".join(unmatched_master),
                "present_in_excelsheet": ", ".join(unmatched_output)
            })
            break

    if not matched:
        unmatched_paths.append({"unmatched_path": path})

# ---------- Create Excel Report ----------
output_file = "comparison_report.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_matched = pd.DataFrame(matched_rows)
    df_unmatched = pd.DataFrame(unmatched_paths)

    df_matched.to_excel(writer, sheet_name="matched paths", index=False)
    df_unmatched.to_excel(writer, sheet_name="unmatched paths", index=False)

# ---------- Apply Header Formatting ----------
wb = load_workbook(output_file)

header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Yellow
header_font = Font(bold=True, color="000000")  # Bold Black

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    for cell in sheet[1]:  # First row (header)
        cell.fill = header_fill
        cell.font = header_font

wb.save(output_file)

print(f" Report generated successfully with formatted headers: {output_file}")
